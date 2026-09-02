import re
import sys
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PDF_PATH    = sys.argv[1] if len(sys.argv) > 1 else "03.pdf"
OUTPUT_PATH = sys.argv[2] if len(sys.argv) > 2 else "colaboradores_sefip.xlsx"

def br_float(s):
    return float(s.strip().replace(".", "").replace(",", "."))

def extrair_texto(pdf_path):
    paginas = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            paginas.append(page.extract_text() or "")
    return paginas

def extrair_meta_pagina(texto):
    # Novo formato real: COMP:02/2015 ou COMP: 02/2015
    comp    = re.search(r"COMP:\s*(\d{2}/\d{4})", texto)
    empresa = re.search(r"EMPRESA:\s*(.+?)\s+INSCRIÇÃO:", texto)
    inscr   = re.search(r"INSCRIÇÃO:\s*([\d./\-]+)", texto)
    return {
        "competencia": comp.group(1).strip()    if comp    else "",
        "empresa":     empresa.group(1).strip() if empresa else "",
        "inscricao":   inscr.group(1).strip()   if inscr   else "",
    }

# Formato real linha A (trabalhador):
# NOME PIS ADMISSÃO CAT [CAT2] [DATA COD_MOV] CBO
# Onde OCOR/DATA são opcionais, e CAT pode aparecer duplicado (ex: "01 05")
PADRAO_A = re.compile(
    r"^(.+?)\s+(\d{3}\.\d{5}\.\d{2}-\d)\s+(\d{2}/\d{2}/\d{4})\s+"
    r"(\d{2}(?:\s+\d{2})?)"               # CAT (simples ou duplo ex: "01 05")
    r"(?:\s+\d{2}/\d{2}/\d{4}\s+\w+)?"   # OCOR DATA/COD_MOV opcional
    r"\s+(\d{5})\s*$"                     # CBO
)

# Formato real linha B (valores):
# REM_SEM_13 | REM_13 | BASE_13_PREV | CONTRIB_SEG | DEPOSITO | JAM
PADRAO_B = re.compile(
    r"^([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$"
)

# Linha Z1/desligamento sem valores (linha A existe mas não há linha B a seguir)
# Detectamos porque a próxima linha começa outro trabalhador ou é TOTAIS

LINHAS_IGNORAR = re.compile(
    r"^(MINISTÉRIO|GFIP|PÁG\s*:|RELAÇÃO|MODALIDADE|EMPRESA:|COMP:|TOMADOR|"
    r"NOME TRABALHADOR|REM SEM|BASE CÁL|TOTAIS|Nº ARQUIVO|Nº DE CONTROLE|"
    r"RESUMO|FGTS|CAT QUANT|TOTAIS:|REMUNERAÇÃO|QUANTIDADE|VALORES DO|"
    r"DATA DE RECOL|DEPÓSITO FGTS|LOGRADOURO|CIDADE|CNAE|FAP:|RAT|"
    r"\d{9,}|[A-Z]{2,}:\s*\d|[A-Z] :|H :|N2:|Q3:|V3:|X :)"
)

def extrair_colaboradores(paginas):
    colaboradores = []

    for texto in paginas:
        meta = extrair_meta_pagina(texto)
        if not meta["competencia"]:
            continue

        linhas = texto.split("\n")
        i = 0
        while i < len(linhas):
            linha = linhas[i]

            ma = PADRAO_A.match(linha)
            if ma:
                nome     = ma.group(1).strip()
                pis      = ma.group(2)
                admissao = ma.group(3)
                cat      = ma.group(4).split()[0]   # só o primeiro CAT
                cbo      = ma.group(5)

                # Verificar se a próxima linha tem valores (pode haver Z1 sem linha B)
                prox = linhas[i + 1] if (i + 1) < len(linhas) else ""
                mb = PADRAO_B.match(prox)

                if mb:
                    rem_sem_13 = br_float(mb.group(1))
                    rem_13     = br_float(mb.group(2))

                    base = {
                        "competencia": meta["competencia"],
                        "empresa":     meta["empresa"],
                        "inscricao":   meta["inscricao"],
                        "nome":        nome,
                        "pis_pasep":   pis,
                        "admissao":    admissao,
                        "cbo":         cbo,
                    }

                    # Linha mensal (sempre, mesmo se zero)
                    colaboradores.append({**base, "tipo": "M", "valor": rem_sem_13})

                    # Linha 13º somente se valor > 0
                    if rem_13 > 0:
                        colaboradores.append({**base, "tipo": "13º", "valor": rem_13})

                    i += 2  # pula linha A e linha B
                else:
                    # Linha Z1 (desligamento sem valores) — ignorar
                    i += 1
            else:
                i += 1

    return colaboradores

def gerar_xlsx(colaboradores, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores SEFIP"

    azul_esc  = "1F3864"
    azul_med  = "2F5496"
    cinza_clr = "D9E1F2"
    verde_tot = "E2EFDA"
    branco    = "FFFFFF"
    COR_13    = "FFF2CC"

    comps = list(dict.fromkeys(c["competencia"] for c in colaboradores))
    cores_comp = ["FFFFFF", "EBF3FB"]

    def fonte(bold=False, size=9, cor=None):
        return Font(name="Arial", bold=bold, size=size, color=cor or "000000")
    def fill(hex_cor):
        return PatternFill("solid", fgColor=hex_cor)

    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esq    = Alignment(horizontal="left",   vertical="center")
    dir_   = Alignment(horizontal="right",  vertical="center")

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"].value     = "GFIP / SEFIP — Relação de Colaboradores"
    ws["A1"].font      = fonte(bold=True, size=14, cor=branco)
    ws["A1"].fill      = fill(azul_esc)
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 24

    comps_str = ", ".join(comps) if comps else ""
    ws.merge_cells("A2:I2")
    ws["A2"].value     = f"Competências: {comps_str}     |     Total de lançamentos: {len(colaboradores)}"
    ws["A2"].font      = fonte(bold=True, size=10, cor=branco)
    ws["A2"].fill      = fill(azul_med)
    ws["A2"].alignment = centro
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 5

    headers = [
        "Competência", "Empresa", "CNPJ",
        "Nome do Trabalhador", "PIS/PASEP/CI", "Admissão",
        "CBO", "Tipo", "Valores (R$)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = fonte(bold=True, cor=branco)
        c.fill = fill(azul_med)
        c.alignment = centro
        c.border = borda
    ws.row_dimensions[4].height = 32

    for idx, colab in enumerate(colaboradores, start=5):
        comp_idx = comps.index(colab["competencia"])

        if colab["tipo"] == "13º":
            bg = fill(COR_13)
        else:
            base_cor = cores_comp[comp_idx % 2]
            bg = fill(base_cor) if base_cor != "FFFFFF" else None

        vals = [
            colab["competencia"], colab["empresa"], colab["inscricao"],
            colab["nome"], colab["pis_pasep"], colab["admissao"],
            colab["cbo"], colab["tipo"], colab["valor"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=col, value=val)
            c.font = fonte()
            c.border = borda
            if bg: c.fill = bg
            if col in (1, 2, 3):
                c.alignment = centro
            elif col == 4:
                c.alignment = esq
            elif col == 9:
                c.number_format = '#,##0.00'
                c.alignment = dir_
            else:
                c.alignment = centro

    tr   = 5 + len(colaboradores)
    last = tr - 1
    ws.merge_cells(f"A{tr}:H{tr}")
    ct = ws[f"A{tr}"]
    ct.value = f"TOTAL GERAL  —  {len(colaboradores)} lançamentos"
    ct.font = fonte(bold=True)
    ct.fill = fill(verde_tot)
    ct.alignment = centro
    ct.border = borda

    c = ws.cell(row=tr, column=9)
    c.value = f"=SUM(I5:I{last})"
    c.font = fonte(bold=True)
    c.fill = fill(verde_tot)
    c.number_format = '#,##0.00'
    c.alignment = dir_
    c.border = borda

    for col, w in enumerate([14, 32, 20, 40, 18, 12, 8, 6, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A5"
    wb.save(output_path)

    print(f"Arquivo salvo: {output_path}")
    print(f"Total lançamentos: {len(colaboradores)}")
    print(f"Competências encontradas: {comps}")

if __name__ == "__main__":
    paginas = extrair_texto(PDF_PATH)
    colaboradores = extrair_colaboradores(paginas)
    gerar_xlsx(colaboradores, output_path=OUTPUT_PATH)
