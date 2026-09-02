"""
=============================================================================
AUTOMAÇÃO - EXTRATO ANALÍTICO FGTS
Conectividade Social ICP V2 - conectividadesocialv2.caixa.gov.br

PRÉ-REQUISITO:
  Abra o Supermium com a flag de debug antes de fazer o login:
    "C:\\Program Files (x86)\\Supermium\\chrome.exe" --remote-debugging-port=9222
  Faça o login normalmente. Depois rode este script.

FLUXO AUTOMATIZADO (por colaborador):
  1. Clicar em "Empregador"
  2. No dropdown, selecionar "Acessar Empresa Outorgante"
  3. Informar o CNPJ e clicar em "Continuar"
  4. No dropdown, selecionar "Solicitar Extrato Analítico do Trabalhador"
  5. Selecionar a Base da Conta
  6. Informar o NIS (PIS/PASEP/NIT)
  7. Clicar em "Confirmar"
  8. Na tela de confirmação, clicar em "Retornar"
  → Repete para o próximo colaborador

Requisitos: pip install -r requirements.txt
=============================================================================
"""

import time
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, WebDriverException
    )
except ImportError:
    print("❌ Selenium não instalado. Execute: pip install -r requirements.txt")
    sys.exit(1)

try:
    import config
except ImportError:
    print("❌ Arquivo config.py não encontrado na pasta do script.")
    sys.exit(1)


# =============================================================================
# UTILITÁRIOS
# =============================================================================

def limpar_numero(valor, zfill: int = 0) -> str:
    """
    Remove tudo que nao for digito.
    O Excel frequentemente converte CNPJs/PIDs para float (ex: 7355965000146.0)
    perdendo zeros a esquerda. Este metodo trata isso corretamente.
    """
    if pd.isna(valor):
        return ""
    # Converte para string e remove notacao cientifica/decimal do Excel
    s = str(valor).strip()
    # Remove parte decimal que o Excel adiciona (ex: "7355965000146.0" -> "7355965000146")
    if s.endswith(".0"):
        s = s[:-2]
    # Remove caracteres nao numericos
    s = re.sub(r"\D", "", s)
    # Repadding com zeros a esquerda se necessario
    if zfill and len(s) < zfill:
        s = s.zfill(zfill)
    return s


def limpar_cnpj(valor) -> str:
    """CNPJ sempre com 14 digitos, preservando zeros a esquerda."""
    return limpar_numero(valor, zfill=14)


def limpar_pis(valor) -> str:
    """PIS/NIS sempre com 11 digitos, preservando zeros a esquerda."""
    return limpar_numero(valor, zfill=11)


def formatar_cnpj(cnpj: str) -> str:
    c = limpar_cnpj(cnpj)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"


def log(msg: str, nivel: str = "INFO"):
    timestamp = datetime.now().strftime("%H:%M:%S")
    icone = {
        "INFO":  "ℹ️ ",
        "OK":    "✅",
        "ERRO":  "❌",
        "AVISO": "⚠️ ",
        "ETAPA": "🔹",
    }.get(nivel, "• ")
    print(f"[{timestamp}] {icone} {msg}")


# =============================================================================
# EXCEÇÃO DE SESSÃO EXPIRADA
# =============================================================================

class SessaoExpiradaError(Exception):
    """Levantada quando o sistema redireciona para tela de login."""
    pass


# =============================================================================
# GERAÇÃO DO ARQUIVO DE LOG
# =============================================================================

def salvar_log(resultados: list):
    if not resultados:
        log("Nenhum resultado para salvar.", "AVISO")
        return

    df = pd.DataFrame(resultados)
    caminho = Path(config.ARQUIVO_LOG)
    df.to_excel(caminho, index=False)

    wb = load_workbook(caminho)
    ws = wb.active

    # Cabeçalho
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Linhas coloridas por status
    verde    = PatternFill("solid", fgColor="C6EFCE")
    vermelho = PatternFill("solid", fgColor="FFC7CE")
    amarelo  = PatternFill("solid", fgColor="FFEB9C")

    for row in ws.iter_rows(min_row=2):
        status = str(row[6].value).upper() if row[6].value else ""
        fill = verde if "SUCESSO" in status else (vermelho if "ERRO" in status or "FALHA" in status else amarelo)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left")

    # Largura das colunas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(caminho)
    log(f"Log salvo em: {caminho.resolve()}", "OK")


# =============================================================================
# CONEXÃO AO SUPERMIUM JÁ ABERTO
# =============================================================================

def conectar_supermium() -> webdriver.Chrome:
    """
    Conecta ao Supermium que já está aberto e logado via remote debugging.
    Não abre nova janela, não faz login.
    """
    porta = getattr(config, "REMOTE_DEBUGGING_PORT", 9222)

    chromedriver_exe = Path(config.CHROMEDRIVER_PATH)
    if not chromedriver_exe.exists():
        log(f"ChromeDriver não encontrado: {chromedriver_exe}", "ERRO")
        log("Verifique CHROMEDRIVER_PATH no config.py.", "ERRO")
        sys.exit(1)

    options = Options()
    options.add_experimental_option("debuggerAddress", f"127.0.0.1:{porta}")

    if config.SUPERMIUM_PATH:
        sup = Path(config.SUPERMIUM_PATH)
        if sup.exists():
            options.binary_location = str(sup)

    service = Service(str(chromedriver_exe))

    try:
        driver = webdriver.Chrome(service=service, options=options)
        driver.implicitly_wait(3)
        log(f"Conectado ao Supermium na porta {porta}.", "OK")
        log(f"Página atual: {driver.current_url}", "INFO")
        return driver
    except WebDriverException as e:
        erro = str(e)
        print()
        print("─" * 60)
        print("  ❌  Não foi possível conectar ao Supermium.")
        print()
        if "cannot connect" in erro.lower() or "connection refused" in erro.lower():
            print(f"  O Supermium não está rodando com a porta {porta} aberta.")
            print()
            print("  SOLUÇÃO: Feche o Supermium e reabra com o comando:")
            print(f'  "{config.SUPERMIUM_PATH}" --remote-debugging-port={porta}')
            print()
            print("  Depois faça o login e rode o script novamente.")
        else:
            print(f"  Detalhe: {erro}")
        print("─" * 60)
        sys.exit(1)


# =============================================================================
# DETECÇÃO DE SESSÃO EXPIRADA
# =============================================================================

def checar_sessao_expirada(driver: webdriver.Chrome):
    """
    Verifica se o sistema redirecionou para tela de login/certificado.
    Se sim, levanta SessaoExpiradaError para interromper a execução.
    """
    try:
        url = driver.current_url.lower()
        corpo = driver.find_element(By.TAG_NAME, "body").text.lower()

        sinais_expiracao = [
            "token", "certificado digital", "autenticação",
            "login", "entrar", "acesso negado", "sessão expirada",
            "sessao expirada",
        ]
        # URLs que indicam tela de login
        urls_login = ["login", "autent", "token", "acesso", "sicscn"]

        if any(p in url for p in urls_login) and "sicse" not in url:
            raise SessaoExpiradaError("URL indica tela de login.")

        if any(s in corpo for s in sinais_expiracao) and "empregador" not in corpo:
            raise SessaoExpiradaError("Conteúdo da página indica sessão expirada.")

    except SessaoExpiradaError:
        raise
    except Exception:
        pass  # Não bloquear por erros de leitura de página


# =============================================================================
# CLASSE DE NAVEGAÇÃO — FLUXO DAS 8 ETAPAS
# =============================================================================

class ConectividadeSocial:
    """
    Fluxo exato por colaborador:

      [TELA INICIAL]
        1. Clicar em "Empregador"

      [TELA SERVIÇOS AO EMPREGADOR]
        2. Dropdown → selecionar "Acessar Empresa Outorgante"
        3. Campo CNPJ → preencher → clicar "Continuar"

      [TELA DE SERVIÇOS DA EMPRESA OUTORGANTE]
        4. Dropdown → selecionar "Solicitar Extrato Analítico do Trabalhador"

      [TELA DO FORMULÁRIO]
        5. Dropdown "Base da Conta" → selecionar estado configurado
        6. Campo NIS → preencher
        7. Clicar em "Confirmar"

      [TELA DE CONFIRMAÇÃO]
        8. Clicar em "Retornar" → volta para tela de serviços
           (próximo colaborador começa no passo 4 se mesma empresa,
            ou no passo 2 se empresa diferente)
    """

    def __init__(self, driver: webdriver.Chrome):
        self.driver = driver
        self.wait   = WebDriverWait(driver, config.TIMEOUT_PAGINA)
        self._cnpj_ativo = None  # Evita re-entrar na outorgante se for a mesma empresa

    def _pausa(self, seg: float = None):
        time.sleep(seg if seg is not None else config.PAUSA_ENTRE_ACOES)

    def _clicar(self, by, valor, descricao="elemento"):
        log(f"Clicando: {descricao}", "ETAPA")
        el = self.wait.until(EC.element_to_be_clickable((by, valor)))
        self._pausa(0.4)
        el.click()
        self._pausa(0.6)
        return el

    def _digitar(self, by, valor, texto, descricao="campo"):
        log(f"Preenchendo: {descricao}", "ETAPA")
        el = self.wait.until(EC.presence_of_element_located((by, valor)))
        el.clear()
        el.send_keys(str(texto))
        self._pausa(0.3)
        return el

    def _disparar_onchange(self, elemento):
        """
        Dispara o evento onchange via JavaScript.
        Necessário porque o sistema da Caixa usa onchange para navegar,
        e o Selenium não dispara eventos nativos ao usar Select().
        """
        try:
            self.driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                elemento
            )
        except Exception:
            pass
        try:
            # Fallback: chama diretamente o handler onchange do formulário
            self.driver.execute_script("""
                var sel = arguments[0];
                if (sel.onchange) { sel.onchange(); }
                else if (sel.form) {
                    var ev = document.createEvent('HTMLEvents');
                    ev.initEvent('change', true, true);
                    sel.dispatchEvent(ev);
                }
            """, elemento)
        except Exception:
            pass

    def _selecionar_no_dropdown_servicos(self, value_exato: str, textos: list, descricao: str):
        """
        Localiza o dropdown sltOpcao, seleciona a opção e dispara o onchange.

        Estratégia:
          1. Tenta select_by_value com o value exato do HTML
          2. Tenta select_by_visible_text com cada texto da lista
          3. Busca parcial no texto ou no value
          4. Dispara onchange manualmente após qualquer seleção bem-sucedida
        """
        select = self._localizar_dropdown_servicos()
        el = select._el  # elemento <select> subjacente

        selecionado = False

        # 1. Por value exato
        if value_exato:
            try:
                select.select_by_value(value_exato)
                log(f"'{descricao}' selecionado por value: {value_exato}", "OK")
                selecionado = True
            except Exception:
                pass

        # 2. Por texto visível
        if not selecionado:
            for texto in textos:
                try:
                    select.select_by_visible_text(texto)
                    log(f"'{descricao}' selecionado por texto: '{texto}'", "OK")
                    selecionado = True
                    break
                except Exception:
                    continue

        # 3. Busca parcial em texto e value
        if not selecionado:
            termos_busca = [t.lower() for t in textos]
            for option in select.options:
                opt_text  = option.text.lower()
                opt_value = (option.get_attribute("value") or "").lower()
                if any(t in opt_text or t in opt_value for t in termos_busca):
                    select.select_by_visible_text(option.text)
                    log(f"'{descricao}' selecionado (parcial): '{option.text}'", "OK")
                    selecionado = True
                    break

        if not selecionado:
            opcoes = [f"'{o.text}' [value={o.get_attribute('value')}]"
                      for o in select.options if o.text.strip()]
            raise Exception(f"'{descricao}' nao encontrado. Opcoes: {opcoes}")

        # 4. DISPARA ONCHANGE — essencial para o sistema da Caixa navegar
        self._disparar_onchange(el)
        self._pausa(4.0)  # Aumentado: sistema da Caixa pode demorar para recarregar a página

    def _selecionar_dropdown(self, by, valor, texto_opcao, descricao="dropdown"):
        """Seleciona opção em <select> genérico por texto exato ou parcial e dispara onchange."""
        log(f"Dropdown '{descricao}': selecionando '{texto_opcao}'", "ETAPA")
        el = self.wait.until(EC.presence_of_element_located((by, valor)))
        select = Select(el)

        selecionado = False
        try:
            select.select_by_visible_text(texto_opcao)
            selecionado = True
        except Exception:
            pass

        if not selecionado:
            for option in select.options:
                if texto_opcao.upper() in option.text.upper():
                    select.select_by_visible_text(option.text)
                    log(f"Opcao encontrada (parcial): '{option.text}'", "INFO")
                    selecionado = True
                    break

        if not selecionado:
            opcoes = [o.text.strip() for o in select.options if o.text.strip()]
            raise Exception(
                f"Opcao '{texto_opcao}' nao encontrada em '{descricao}'.\n"
                f"Opcoes disponiveis: {opcoes}"
            )

        self._disparar_onchange(el)
        self._pausa(1)

    def _localizar_dropdown_servicos(self) -> Select:
        """
        Localiza o dropdown principal de serviços da página.
        HTML real do sistema: <select name="sltOpcao" class="txtcentral" ...>
        Confirmado via DevTools (imagens do usuário).
        """
        seletores = [
            (By.NAME,  "sltOpcao"),                      # name real — confirmado no DevTools
            (By.XPATH, "//select[@class='txtcentral']"), # class real — confirmada no DevTools
            (By.XPATH, "//select[contains(@name,'Opcao') or contains(@name,'opcao')]"),
            (By.XPATH, "//select"),                      # fallback: primeiro select da página
        ]
        for by, valor in seletores:
            try:
                el = WebDriverWait(self.driver, 8).until(
                    EC.presence_of_element_located((by, valor))
                )
                s = Select(el)
                # Confirma que é o dropdown certo verificando opções de serviço
                textos = [o.text for o in s.options]
                if any("Extrato" in t or "Outorgante" in t or "Trabalhador" in t for t in textos):
                    log(f"Dropdown sltOpcao localizado ({len(s.options)} opcoes).", "INFO")
                    return s
            except Exception:
                continue
        raise Exception("Dropdown de servicos (sltOpcao) nao encontrado na pagina.")

    # ------------------------------------------------------------------
    # ETAPA 1 — Clicar em "Empregador" na tela inicial
    # ------------------------------------------------------------------
    def etapa1_clicar_empregador(self):
        """
        Na tela inicial do Conectividade Social V2, clica no botão/link 'Empregador'.
        O sistema redireciona para sicse.caixa.gov.br (Serviços ao Empregador).
        """
        log("ETAPA 1 — Clicando em 'Empregador'", "ETAPA")
        checar_sessao_expirada(self.driver)
        self._clicar(
            By.XPATH,
            "//a[normalize-space(text())='Empregador'] "
            "| //button[normalize-space(text())='Empregador'] "
            "| //input[@value='Empregador'] "
            "| //a[contains(@href,'Empregador')] "
            "| //*[normalize-space(text())='Empregador' and (self::a or self::button or self::input)]",
            "botão Empregador"
        )
        self._pausa(2)

    # ------------------------------------------------------------------
    # ETAPA 2 — Dropdown → "Acessar Empresa Outorgante"
    # ------------------------------------------------------------------
    def etapa2_acessar_empresa_outorgante(self):
        """
        Seleciona "Acessar Empresa Outorgante" no dropdown sltOpcao e dispara onchange.
        value real (DevTools): "AcessarOutorgante|AcessarOutorgante.Verificar"
        O onchange chama funcionalidades() que faz a navegação.
        """
        log("ETAPA 2 — Selecionando 'Acessar Empresa Outorgante'", "ETAPA")
        checar_sessao_expirada(self.driver)
        self._selecionar_no_dropdown_servicos(
            value_exato="AcessarOutorgante|AcessarOutorgante.Verificar",
            textos=["Acessar Empresa Outorgante", "Empresa Outorgante"],
            descricao="Acessar Empresa Outorgante"
        )

    # ------------------------------------------------------------------
    # ETAPA 3 — Preencher CNPJ e clicar em "Continuar"
    # ------------------------------------------------------------------
    def etapa3_informar_cnpj(self, cnpj: str):
        cnpj_limpo = limpar_numero(cnpj)
        log(f"ETAPA 3 — Informando CNPJ: {formatar_cnpj(cnpj_limpo)}", "ETAPA")
        checar_sessao_expirada(self.driver)

        self._digitar(
            By.XPATH,
            "//input[contains(translate(@id,'cnpj','CNPJ'),'CNPJ') or "
            "contains(translate(@name,'cnpj','CNPJ'),'CNPJ')]",
            cnpj_limpo,
            "campo CNPJ"
        )

        # O botão CONTINUAR é uma <a href="javascript:subm_verificar_empresa()">
        # com imagem interna — confirmado via DevTools. Não é input nem button.
        self._clicar(
            By.XPATH,
            "//a[contains(@href,'subm_verificar_empresa')] | "
            "//a[contains(@href,'verificar_empresa')] | "
            "//img[contains(@src,'botao_continuar')]/parent::a | "
            "//img[@name='cmdCont']/parent::a | "
            "//input[@value='CONTINUAR' or @value='Continuar'] | "
            "//button[contains(translate(text(),'continuar','CONTINUAR'),'CONTINUAR')]",
            "botão CONTINUAR (subm_verificar_empresa)"
        )
        self._pausa(2)
        self._cnpj_ativo = cnpj_limpo
        log(f"CNPJ {formatar_cnpj(cnpj_limpo)} confirmado.", "OK")

    # ------------------------------------------------------------------
    # ETAPA 4 — Dropdown → "Solicitar Extrato Analítico do Trabalhador"
    # ------------------------------------------------------------------
    def etapa4_selecionar_extrato_analitico(self):
        """
        Seleciona "Solicitar Extrato Analítico do Trabalhador" no dropdown sltOpcao e dispara onchange.
        value real (DevTools): "300|ExtratoAnaliticoTrabalhador.Solicitar"
        O onchange chama funcionalidades() que faz a navegação.
        """
        log("ETAPA 4 — Selecionando 'Solicitar Extrato Analitico do Trabalhador'", "ETAPA")
        checar_sessao_expirada(self.driver)
        self._selecionar_no_dropdown_servicos(
            value_exato="300|ExtratoAnaliticoTrabalhador.Solicitar",
            textos=[
                "Solicitar Extrato Analítico do Trabalhador",
                "Solicitar Extrato Analitico do Trabalhador",
            ],
            descricao="Solicitar Extrato Analitico do Trabalhador"
        )

    # ------------------------------------------------------------------
    # ETAPA 5 — Selecionar Base da Conta
    # ------------------------------------------------------------------
    def etapa5_selecionar_base_conta(self):
        log(f"ETAPA 5 — Selecionando Base da Conta: '{config.BASE_DA_CONTA}'", "ETAPA")
        checar_sessao_expirada(self.driver)

        # DevTools confirmado: <select name="sltRegiao" class="txtcentral" size="1">
        # IMPORTANTE: aguarda sltRegiao aparecer no DOM antes de qualquer ação.
        # Após a etapa 4, o sistema navega/recarrega a página — sltRegiao só existe
        # no formulário do Extrato Analítico, não na tela anterior (que tem sltOpcao).
        # Sem este wait, _selecionar_dropdown encontra sltOpcao (mesma class) e falha.
        log("Aguardando formulário do Extrato Analítico carregar...", "INFO")
        try:
            self.wait.until(
                EC.presence_of_element_located((By.NAME, "sltRegiao"))
            )
            log("Formulário carregado — sltRegiao presente.", "OK")
        except TimeoutException:
            # Fallback: aguarda pelo campo txtPIS como sinal alternativo de carregamento
            log("sltRegiao não detectado via NAME, tentando via txtPIS...", "AVISO")
            try:
                self.wait.until(
                    EC.presence_of_element_located((By.ID, "txtPIS"))
                )
                log("Formulário carregado — txtPIS presente.", "OK")
            except TimeoutException:
                raise Exception(
                    "Formulário do Extrato Analítico não carregou após etapa 4. "
                    "Verifique se o onchange da etapa 4 está navegando corretamente."
                )

        # DevTools confirmado — values reais do sltRegiao:
        #   "GOD" → GO-GOIAS
        #   "BHD" → BH-BELO HORIZONTE
        #   "SPD" → SP-SAO PAULO
        # config.BASE_DA_CONTA usa o texto (ex: "SP-SAO PAULO") → convertemos para value
        MAPA_BASE_CONTA = {
            "GO":  "GOD", "GOIAS":       "GOD", "GO-GOIAS":        "GOD",
            "BH":  "BHD", "BELO":        "BHD", "BH-BELO":         "BHD", "BH-BELO HORIZONTE": "BHD",
            "SP":  "SPD", "SAO PAULO":   "SPD", "SP-SAO PAULO":    "SPD",
        }
        base_upper = config.BASE_DA_CONTA.strip().upper()
        value_base = MAPA_BASE_CONTA.get(base_upper)

        # Se não achou no mapa, tenta match parcial
        if not value_base:
            for chave, val in MAPA_BASE_CONTA.items():
                if chave in base_upper or base_upper in chave:
                    value_base = val
                    break

        el = self.wait.until(EC.presence_of_element_located((By.NAME, "sltRegiao")))
        select = Select(el)

        selecionado = False
        if value_base:
            try:
                select.select_by_value(value_base)
                log(f"Base da Conta selecionada por value: '{value_base}' ({config.BASE_DA_CONTA})", "OK")
                selecionado = True
            except Exception:
                pass

        if not selecionado:
            # Fallback: busca parcial no texto visível
            for option in select.options:
                if base_upper in option.text.strip().upper():
                    select.select_by_visible_text(option.text)
                    log(f"Base da Conta selecionada por texto: '{option.text}'", "OK")
                    selecionado = True
                    break

        if not selecionado:
            opcoes = [(o.get_attribute("value"), o.text.strip()) for o in select.options if o.text.strip()]
            raise Exception(
                f"Base da Conta '{config.BASE_DA_CONTA}' nao encontrada em sltRegiao.\n"
                f"Opcoes disponiveis (value → texto): {opcoes}\n"
                f"Ajuste BASE_DA_CONTA no config.py para um dos valores acima."
            )
        self._pausa(0.5)

    # ------------------------------------------------------------------
    # ETAPA 6 — Informar NIS (PIS/PASEP/NIT)
    # ------------------------------------------------------------------
    def etapa6_informar_nis(self, nis: str):
        nis_limpo = limpar_numero(nis)
        log(f"ETAPA 6 — Informando NIS: {nis_limpo}", "ETAPA")
        checar_sessao_expirada(self.driver)

        # DevTools confirmado: <input id="txtPIS" name="txtPIS" type="Text" size="12" maxlength="11">
        self._digitar(
            By.XPATH,
            "//input[@id='txtPIS'] | "
            "//input[@name='txtPIS'] | "
            "//input[contains(@id,'PIS') or contains(@name,'PIS')] | "
            "//input[contains(@id,'NIS') or contains(@name,'NIS')]",
            nis_limpo,
            "campo NIS/PIS (txtPIS)"
        )

    # ------------------------------------------------------------------
    # ETAPA 7 — Clicar em "Confirmar"
    # ------------------------------------------------------------------
    def etapa7_confirmar(self):
        log("ETAPA 7 — Clicando em 'Confirmar'", "ETAPA")
        checar_sessao_expirada(self.driver)

        # DevTools confirmado: <a href="javascript:subm_extrato_analitico_trabalhador()">
        #   <img src="/sicse/images/botao_confirmar.gif" name="cmdCont" ...>
        self._clicar(
            By.XPATH,
            "//a[contains(@href,'subm_extrato_analitico_trabalhador')] | "
            "//img[@name='cmdCont']/parent::a | "
            "//img[contains(@src,'botao_confirmar')]/parent::a | "
            "//a[contains(@href,'confirmar')] | "
            "//input[@value='CONFIRMAR' or @value='Confirmar'] | "
            "//button[contains(translate(text(),'confirmar','CONFIRMAR'),'CONFIRMAR')]",
            "botão Confirmar (subm_extrato_analitico_trabalhador)"
        )
        self._pausa(2)

    # ------------------------------------------------------------------
    # ETAPA 8 — Verificar confirmação e clicar em "Retornar"
    # ------------------------------------------------------------------
    def etapa8_verificar_e_retornar(self) -> tuple[bool, str]:
        log("ETAPA 8 — Verificando confirmação e retornando", "ETAPA")
        checar_sessao_expirada(self.driver)

        corpo = self.driver.find_element(By.TAG_NAME, "body").text
        corpo_lower = corpo.lower()

        if "efetuada com sucesso" in corpo_lower or "solicitação efetuada" in corpo_lower:
            sucesso, mensagem = True, "Solicitação efetuada com sucesso"
        elif "erro" in corpo_lower or "inválido" in corpo_lower or "invalido" in corpo_lower:
            trecho = corpo[:300].replace("\n", " ")
            sucesso, mensagem = False, f"Erro retornado pelo sistema: {trecho}"
        else:
            trecho = corpo[:300].replace("\n", " ")
            sucesso, mensagem = False, f"Resposta indefinida: {trecho}"

        # Clicar em Retornar independente do resultado
        try:
            self._clicar(
                By.XPATH,
                "//a[contains(@href,'Principal.Visualizar') or contains(@href,'Retornar') or contains(@href,'retornar')] | "
                "//img[contains(@src,'botao_retornar') or contains(@src,'retornar')]/parent::a | "
                "//input[@value='RETORNAR' or @value='Retornar'] | "
                "//button[contains(translate(text(),'retornar','RETORNAR'),'RETORNAR')]",
                "botão Retornar"
            )
            self._pausa(1.5)
        except Exception:
            log("Botão 'Retornar' não encontrado — continuando.", "AVISO")

        return sucesso, mensagem

    # ------------------------------------------------------------------
    # FLUXO COMPLETO POR COLABORADOR
    # ------------------------------------------------------------------
    def processar_colaborador(self, cnpj: str, nis: str) -> tuple[bool, str]:
        """
        Executa as etapas 2 a 8 para um colaborador.
        A etapa 1 (clicar Empregador) é feita uma única vez antes do loop.
        Se o CNPJ for o mesmo do colaborador anterior, pula etapas 2 e 3.
        """
        cnpj_limpo = limpar_numero(cnpj)
        mesma_empresa = (self._cnpj_ativo == cnpj_limpo)

        if not mesma_empresa:
            self.etapa2_acessar_empresa_outorgante()
            self.etapa3_informar_cnpj(cnpj_limpo)

        self.etapa4_selecionar_extrato_analitico()
        self.etapa5_selecionar_base_conta()
        self.etapa6_informar_nis(nis)
        self.etapa7_confirmar()
        return self.etapa8_verificar_e_retornar()


# =============================================================================
# LEITURA DA PLANILHA
# =============================================================================

def carregar_colaboradores() -> pd.DataFrame:
    caminho = Path(config.ARQUIVO_EXCEL)
    if not caminho.exists():
        log(f"Arquivo '{caminho}' nao encontrado.", "ERRO")
        log(f"Pasta atual: {Path.cwd()}", "INFO")
        sys.exit(1)

    df = pd.read_excel(caminho, dtype=str)

    # Mapeia colunas da planilha para os nomes do config,
    # ignorando maiusculas, acentos e espacos extras
    import unicodedata

    def normalizar(texto: str) -> str:
        """Remove acentos, converte para minusculo e elimina espacos extras."""
        texto = str(texto).strip().lower()
        texto = unicodedata.normalize("NFD", texto)
        texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
        return texto

    # Cria mapa: nome_normalizado -> nome_original_da_planilha
    mapa_colunas = {normalizar(c): c for c in df.columns}

    # Colunas que precisamos do config
    colunas_config = {
        "empresa":       config.COL_EMPRESA,
        "cnpj":          config.COL_CNPJ_OUTORG,
        "nome":          config.COL_NOME,
        "pis":           config.COL_PIS,
        "cpf":           config.COL_CPF,
        "admissao":      config.COL_DATA_ADMISSAO,
        "data demissao": config.COL_DATA_DEMISSAO,
    }

    # Renomeia as colunas da planilha para os nomes padrao do config
    renomear = {}
    for chave_norm, nome_config in colunas_config.items():
        nome_config_norm = normalizar(nome_config)
        if nome_config_norm in mapa_colunas:
            col_original = mapa_colunas[nome_config_norm]
            if col_original != nome_config:
                renomear[col_original] = nome_config

    if renomear:
        df = df.rename(columns=renomear)

    log(f"Planilha carregada: {len(df)} colaboradores.", "OK")
    log(f"Colunas encontradas: {list(df.columns)}", "INFO")
    return df


# =============================================================================
# PROCESSAMENTO PRINCIPAL
# =============================================================================

def processar_colaboradores():
    df = carregar_colaboradores()
    resultados = []
    driver  = None
    sistema = None

    try:
        driver  = conectar_supermium()
        sistema = ConectividadeSocial(driver)

        # Etapa 1 — feita uma única vez
        sistema.etapa1_clicar_empregador()

        total = len(df)
        for idx, row in df.iterrows():
            numero  = idx + 1
            nome    = str(row.get(config.COL_NOME, "")).strip()
            pis     = limpar_pis(row.get(config.COL_PIS, ""))
            cnpj    = limpar_cnpj(row.get(config.COL_CNPJ_OUTORG, ""))
            empresa = str(row.get(config.COL_EMPRESA, "")).strip()

            log(f"{'─'*55}", "INFO")
            log(f"Colaborador {numero}/{total}: {nome} | PIS: {pis} | Empresa: {empresa}", "INFO")

            resultado = {
                "empresa":         empresa,
                "cnpj_outorgante": formatar_cnpj(cnpj) if cnpj else "",
                "nome":            nome,
                "pis":             pis,
                "cpf":             limpar_numero(row.get(config.COL_CPF, ""), zfill=11),
                "data_admissao":   str(row.get(config.COL_DATA_ADMISSAO, "")),
                "status":          "",
                "mensagem":        "",
                "data_hora":       datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            }

            # Validações básicas
            if not pis:
                resultado.update({"status": "ERRO", "mensagem": "PIS/NIS não informado"})
                resultados.append(resultado)
                log(f"Pulando {nome}: PIS não informado.", "AVISO")
                continue

            if not cnpj:
                resultado.update({"status": "ERRO", "mensagem": "CNPJ não informado"})
                resultados.append(resultado)
                log(f"Pulando {nome}: CNPJ não informado.", "AVISO")
                continue

            try:
                sucesso, mensagem = sistema.processar_colaborador(cnpj, pis)
                resultado["status"]   = "SUCESSO" if sucesso else "FALHA"
                resultado["mensagem"] = mensagem
                log(f"{nome}: {mensagem}", "OK" if sucesso else "ERRO")

            except SessaoExpiradaError as e:
                # Sessão expirou — salva log e para tudo
                resultado.update({"status": "ERRO", "mensagem": f"Sessão expirada: {e}"})
                resultados.append(resultado)
                salvar_log(resultados)
                print()
                print("=" * 60)
                print("  ⚠️  SESSÃO EXPIRADA — AUTOMAÇÃO PAUSADA")
                print()
                print(f"  Último colaborador processado: {nome}")
                print(f"  Colaboradores restantes: {total - numero}")
                print()
                print("  O que fazer:")
                print("  1. Faça o login novamente no Supermium")
                print("  2. Volte para a tela inicial do sistema")
                print(f"  3. Edite a planilha e remova os {numero} já processados")
                print("  4. Rode o script novamente")
                print()
                print(f"  Log parcial salvo em: {config.ARQUIVO_LOG}")
                print("=" * 60)
                return  # Interrompe sem fechar o navegador

            except Exception as e:
                resultado.update({"status": "ERRO", "mensagem": str(e)})
                log(f"Erro em {nome}: {e}", "ERRO")
                # Tenta recuperar: volta à tela inicial e recomeça do passo 1
                try:
                    log("Tentando recuperar sessão...", "AVISO")
                    sistema._cnpj_ativo = None
                    driver.get(config.URL_SISTEMA)
                    time.sleep(3)
                    checar_sessao_expirada(driver)
                    sistema.etapa1_clicar_empregador()
                    log("Sessão recuperada. Continuando.", "OK")
                except SessaoExpiradaError:
                    resultado.update({"status": "ERRO", "mensagem": "Sessão expirada durante recuperação"})
                    resultados.append(resultado)
                    salvar_log(resultados)
                    print()
                    print("=" * 60)
                    print("  ⚠️  SESSÃO EXPIRADA — AUTOMAÇÃO PAUSADA")
                    print(f"  Log parcial salvo em: {config.ARQUIVO_LOG}")
                    print("=" * 60)
                    return
                except Exception as e2:
                    log(f"Não foi possível recuperar: {e2}", "ERRO")

            resultados.append(resultado)
            time.sleep(config.PAUSA_ENTRE_COLABORADORES)

    except KeyboardInterrupt:
        log("Execução interrompida pelo usuário (Ctrl+C).", "AVISO")
    except Exception as e:
        log(f"Erro crítico: {e}", "ERRO")
        import traceback
        traceback.print_exc()
    finally:
        salvar_log(resultados)
        log("Navegador mantido aberto.", "OK")

    # Resumo final
    total_res = len(resultados)
    sucessos  = sum(1 for r in resultados if r["status"] == "SUCESSO")
    erros     = sum(1 for r in resultados if r["status"] in ("ERRO", "FALHA"))
    print()
    print("=" * 60)
    print("  RESUMO DA EXECUÇÃO")
    print(f"  Total processados : {total_res}")
    print(f"  ✅ Sucesso         : {sucessos}")
    print(f"  ❌ Erros/Falhas    : {erros}")
    print(f"  📄 Log salvo em   : {config.ARQUIVO_LOG}")
    print("=" * 60)


# =============================================================================
# PONTO DE ENTRADA
# =============================================================================

if __name__ == "__main__":
    porta = getattr(config, "REMOTE_DEBUGGING_PORT", 9222)

    print()
    print("=" * 60)
    print("  EXTRATO ANALÍTICO FGTS — CONECTIVIDADE SOCIAL V2")
    print(f"  Base da Conta  : {config.BASE_DA_CONTA}")
    print(f"  Planilha       : {config.ARQUIVO_EXCEL}")
    print(f"  Porta de debug : {porta}")
    print("=" * 60)
    print()
    print("  Antes de continuar, confirme:")
    print("  ✔  Supermium aberto com --remote-debugging-port=" + str(porta))
    print("  ✔  Login já realizado no sistema")
    print("  ✔  Tela inicial do Conectividade Social visível")
    print()
    input("  Pressione ENTER para iniciar a automação...")
    print()
    processar_colaboradores()
