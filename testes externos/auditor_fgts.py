"""
auditor_fgts.py
---------------
Auditor de Extrato Analítico FGTS (Caixa Econômica Federal)
Lê arquivo XLSX no padrão CEF e gera relatório analítico + sintético.

Uso:
    python auditor_fgts.py <arquivo_entrada.xlsx> [arquivo_saida.xlsx]

Se arquivo_saida não for informado, gera: auditoria_fgts_AAAAMMDD_HHMMSS.xlsx
"""

import sys
import re
import os
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────

MESES = {
    "JANEIRO": "01", "FEVEREIRO": "02", "MARCO": "03", "MARÇO": "03",
    "ABRIL": "04", "MAIO": "05", "JUNHO": "06", "JULHO": "07",
    "AGOSTO": "08", "SETEMBRO": "09", "OUTUBRO": "10",
    "NOVEMBRO": "11", "DEZEMBRO": "12",
}

RE_LANCAMENTO  = re.compile(r"^(\d{2}/\d{2}/\d{4})\s{2,}(.+?)\s{2,}(-?[\d.,]+)\s*$")
RE_DEP_NORMAL  = re.compile(r"^DEPOSITO\s+([A-ZÇÃÕÁÉÍÓÚ]+)/(\d{4})$", re.IGNORECASE)
RE_DEP_ATRASO  = re.compile(r"^DEPOSITO\s+EM\s+ATRASO\s+([A-ZÇÃÕÁÉÍÓÚ]+)/(\d{4})$", re.IGNORECASE)

# Cores
COR_HEADER       = "1F3864"
COR_HEADER_FONT  = "FFFFFF"
COR_ATRASO       = "FFF2CC"
COR_FALTANTE     = "FCE4D6"   # laranja claro — competência faltante
COR_FALTANTE_FNT = "833C00"
COR_NORMAL_PAR   = "EBF1F8"
COR_NORMAL_IMPAR = "FFFFFF"
COR_SINTETICO    = "E2EFDA"
COR_SIN_HEADER   = "375623"
COR_SIN_ALERTA   = "FF0000"   # vermelho para colaboradores com faltantes

THIN  = Side(border_style="thin", color="BFBFBF")
BORDA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATA_EXTRATO = date.today()   # referência para colaboradores ainda ativos


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def normalizar_valor(texto: str) -> float:
    return float(texto.strip().replace(".", "").replace(",", "."))


def formatar_cnpj(raw: str) -> str:
    d = re.sub(r"\D", "", raw)
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return raw


def competencia_para_mmaaaa(mes_str: str, ano_str: str) -> str:
    num = MESES.get(mes_str.upper().strip())
    return f"{num}/{ano_str}" if num else f"??/{ano_str}"


def mmaaaa_para_date(comp: str) -> date | None:
    """'08/2019' → date(2019, 8, 1)"""
    try:
        m, a = comp.split("/")
        return date(int(a), int(m), 1)
    except Exception:
        return None


def ddmmaaaa_para_date(s: str) -> date | None:
    try:
        d, m, a = s.split("/")
        return date(int(a), int(m), int(d))
    except Exception:
        return None


def extrair_data_afastamento(linha: str) -> str | None:
    datas = re.findall(r"\b(\d{2}/\d{2}/\d{4})\b", linha)
    if len(datas) >= 3:
        return datas[2] if datas[2] != "00/00/0000" else None
    return None


def gerar_competencias_esperadas(adm: str, dem: str | None) -> list[str]:
    """
    Retorna lista de competências MM/AAAA que deveriam ter depósito,
    do mês de admissão até o mês de demissão (ou mês anterior ao atual
    se ainda ativo).

    Regra: o depósito de competência MM/AAAA deve existir se o colaborador
    estava ativo em qualquer dia daquele mês.
    """
    inicio = ddmmaaaa_para_date(adm)
    if not inicio:
        return []

    if dem:
        fim_date = ddmmaaaa_para_date(dem)
        # Último mês com obrigação = mês da demissão (se demitido no meio
        # do mês, o FGTS do mês inteiro ainda é devido)
        fim = date(fim_date.year, fim_date.month, 1) if fim_date else DATA_EXTRATO.replace(day=1)
    else:
        # Ativo: última competência esperada = mês anterior ao atual
        # (o mês corrente ainda pode não ter vencido)
        ref = DATA_EXTRATO.replace(day=1) - relativedelta(months=1)
        fim = ref

    competencias = []
    cursor = date(inicio.year, inicio.month, 1)
    while cursor <= fim:
        competencias.append(f"{cursor.month:02d}/{cursor.year}")
        cursor += relativedelta(months=1)
    return competencias


# ──────────────────────────────────────────────────────────────────────────────
# PARSER PRINCIPAL
# ──────────────────────────────────────────────────────────────────────────────

def parse_extrato(caminho: str) -> list[dict]:
    wb = load_workbook(caminho, read_only=True)
    ws = wb.active

    linhas = [str(row[0]).rstrip() for row in ws.iter_rows(values_only=True)
              if row and row[0] is not None]

    registros = []
    ctx = {}
    aguardando_pis = aguardando_afastamento = aguardando_empregador = False
    em_lancamentos = False

    i = 0
    while i < len(linhas):
        linha = linhas[i]

        if "NOME DO TRABALHADOR" in linha and "NUM.CONTA" in linha:
            if i + 1 < len(linhas):
                linha_nome = linhas[i + 1]
                nome = linha_nome[:43].strip()
                pag_m = re.search(r"(\d+)/\s*(\d+)\s*$", linha_nome)
                pag_atual = int(pag_m.group(1)) if pag_m else 1
                if pag_atual == 1:
                    ctx = {"nome_trabalhador": nome}
                    em_lancamentos = False
            aguardando_pis = True
            i += 2
            continue

        if aguardando_pis and "PIS/PASEP" in linha and "DTA.ADM" in linha:
            if i + 1 < len(linhas):
                dados = linhas[i + 1]
                pis_m = re.match(r"^(\d{11})", dados)
                adm_m = re.search(r"(\d{2}/\d{2}/\d{4})", dados)
                if pis_m: ctx["pis"] = pis_m.group(1)
                if adm_m: ctx["data_admissao"] = adm_m.group(1)
            aguardando_pis = False
            aguardando_afastamento = True
            i += 2
            continue

        if aguardando_afastamento and "DATA DE AFAST." in linha:
            if i + 1 < len(linhas):
                ctx["data_demissao"] = extrair_data_afastamento(linhas[i + 1]) or ""
            aguardando_afastamento = False
            aguardando_empregador = True
            i += 2
            continue

        if aguardando_empregador and "NOME DO EMPREGADOR" in linha:
            if i + 1 < len(linhas):
                emp_linha = linhas[i + 1]
                cnpj_m = re.search(r"(\d{14})\s*$", emp_linha)
                ctx["nome_empresa"] = emp_linha[:52].strip()
                ctx["cnpj"] = formatar_cnpj(cnpj_m.group(1)) if cnpj_m else ""
            aguardando_empregador = False
            i += 2
            continue

        if "HISTORICO DOS LANCAMENTOS" in linha:
            em_lancamentos = True
            i += 1
            continue

        if em_lancamentos and (
            "SALDO A TRANSPORTAR" in linha
            or (linha.startswith("SALDO ") and "DEPOSITO" in linha)
        ):
            em_lancamentos = False
            i += 1
            continue

        if em_lancamentos:
            m = RE_LANCAMENTO.match(linha)
            if m:
                data_lanc = m.group(1)
                descricao = m.group(2).strip()
                valor_txt = m.group(3)
                dep_normal = RE_DEP_NORMAL.match(descricao)
                dep_atraso = RE_DEP_ATRASO.match(descricao)
                if dep_normal or dep_atraso:
                    mo = dep_normal or dep_atraso
                    competencia = competencia_para_mmaaaa(mo.group(1), mo.group(2))
                    em_atraso = bool(dep_atraso)
                    try:
                        valor = normalizar_valor(valor_txt)
                    except ValueError:
                        valor = 0.0
                    registros.append({
                        "nome_empresa":     ctx.get("nome_empresa", ""),
                        "cnpj":             ctx.get("cnpj", ""),
                        "nome_trabalhador": ctx.get("nome_trabalhador", ""),
                        "pis":              ctx.get("pis", ""),
                        "data_admissao":    ctx.get("data_admissao", ""),
                        "data_demissao":    ctx.get("data_demissao", ""),
                        "data_lancamento":  data_lanc,
                        "valor":            valor,
                        "competencia":      competencia,
                        "em_atraso":        em_atraso,
                        "descricao_raw":    descricao,
                    })
        i += 1

    return registros


# ──────────────────────────────────────────────────────────────────────────────
# ANÁLISE DE COMPETÊNCIAS FALTANTES
# ──────────────────────────────────────────────────────────────────────────────

def calcular_faltantes(registros: list[dict]) -> dict:
    """
    Para cada (empresa, colaborador) calcula:
    - competências esperadas (período admissão → demissão/hoje)
    - competências depositadas (cobertas por qualquer depósito, regular ou em atraso)
    - competências FALTANTES = esperadas - depositadas

    Retorna dict: chave (empresa, colaborador) → {
        "faltantes": [...],
        "esperadas": [...],
        "depositadas": set(),
        "data_admissao": str,
        "data_demissao": str,
        "cnpj": str,
        "pis": str,
    }
    """
    # Agrupa competências depositadas por colaborador
    depositadas_map: dict[tuple, set] = defaultdict(set)
    meta: dict[tuple, dict] = {}

    for r in registros:
        chave = (r["nome_empresa"], r["nome_trabalhador"])
        depositadas_map[chave].add(r["competencia"])
        if chave not in meta:
            meta[chave] = {
                "data_admissao": r["data_admissao"],
                "data_demissao": r["data_demissao"],
                "cnpj": r["cnpj"],
                "pis":  r["pis"],
            }
        else:
            # Mantém data_demissao mais recente capturada (caso multi-vínculo)
            if r["data_demissao"]:
                meta[chave]["data_demissao"] = r["data_demissao"]

    resultado = {}
    for chave, depositadas in depositadas_map.items():
        m = meta[chave]
        esperadas = gerar_competencias_esperadas(
            m["data_admissao"], m["data_demissao"] or None
        )
        faltantes = sorted(
            [c for c in esperadas if c not in depositadas],
            key=lambda x: (int(x.split("/")[1]), int(x.split("/")[0]))
        )
        resultado[chave] = {
            "faltantes":    faltantes,
            "esperadas":    esperadas,
            "depositadas":  depositadas,
            "data_admissao": m["data_admissao"],
            "data_demissao": m["data_demissao"],
            "cnpj": m["cnpj"],
            "pis":  m["pis"],
        }

    return resultado


# ──────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────────────────────────────────────

def estilo_header(cell, cor_fundo=COR_HEADER, cor_fonte=COR_HEADER_FONT):
    cell.font      = Font(bold=True, color=cor_fonte, size=10)
    cell.fill      = PatternFill("solid", start_color=cor_fundo)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDA


def estilo_celula(cell, cor_fundo, bold=False, align="left", font_color="000000"):
    cell.fill      = PatternFill("solid", start_color=cor_fundo)
    cell.font      = Font(bold=bold, size=9, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BORDA


def auto_width(ws, extra=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 70)


# ──────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DO XLSX
# ──────────────────────────────────────────────────────────────────────────────

def gerar_xlsx(registros: list[dict], caminho_saida: str):
    wb = Workbook()

    analise_faltantes = calcular_faltantes(registros)

    # ── ABA ANALÍTICA ─────────────────────────────────────────────────────────
    ws_ana = wb.active
    ws_ana.title = "Analítico"
    ws_ana.freeze_panes = "A2"

    col_headers = [
        "Empresa", "CNPJ", "Colaborador", "PIS/PASEP",
        "Data Admissão", "Data Demissão",
        "Data Lançamento", "Competência", "Valor (R$)",
        "Em Atraso?", "Observação",
    ]
    for c, h in enumerate(col_headers, 1):
        estilo_header(ws_ana.cell(row=1, column=c, value=h))
    ws_ana.row_dimensions[1].height = 28

    for idx, r in enumerate(registros, 2):
        par = idx % 2 == 0
        cor_row = COR_ATRASO if r["em_atraso"] else (COR_NORMAL_PAR if par else COR_NORMAL_IMPAR)
        valores = [
            r["nome_empresa"], r["cnpj"], r["nome_trabalhador"], r["pis"],
            r["data_admissao"], r["data_demissao"],
            r["data_lancamento"], r["competencia"], r["valor"],
            "SIM ⚠️" if r["em_atraso"] else "NÃO",
            "Depósito em atraso – valor pode divergir dos 8% devidos" if r["em_atraso"] else "",
        ]
        for c, val in enumerate(valores, 1):
            cell = ws_ana.cell(row=idx, column=c, value=val)
            align = "right" if c == 9 else ("center" if c in (5, 6, 7, 8, 10) else "left")
            estilo_celula(cell, cor_row, bold=(c == 9), align=align)
            if c == 9:
                cell.number_format = '#,##0.00'

    auto_width(ws_ana)

    # ── ABA COMPETÊNCIAS FALTANTES ────────────────────────────────────────────
    ws_falt = wb.create_sheet("Competências Faltantes")
    ws_falt.freeze_panes = "A2"

    falt_headers = [
        "Empresa", "CNPJ", "Colaborador", "PIS/PASEP",
        "Data Admissão", "Data Demissão",
        "Competências Esperadas", "Competências Depositadas",
        "Qtd Faltantes", "Competências Faltantes",
        "Situação",
    ]
    for c, h in enumerate(falt_headers, 1):
        estilo_header(ws_falt.cell(row=1, column=c, value=h), "833C00")
    ws_falt.row_dimensions[1].height = 28

    # Ordena: colaboradores COM faltantes primeiro, depois sem
    chaves_ordenadas = sorted(
        analise_faltantes.keys(),
        key=lambda k: (len(analise_faltantes[k]["faltantes"]) == 0, k[1])
    )

    for idx, chave in enumerate(chaves_ordenadas, 2):
        empresa, colaborador = chave
        d = analise_faltantes[chave]
        tem_faltante = len(d["faltantes"]) > 0

        if tem_faltante:
            cor_row  = COR_FALTANTE
            fcor     = COR_FALTANTE_FNT
            situacao = f"❌ {len(d['faltantes'])} competência(s) sem depósito"
        else:
            cor_row  = COR_SINTETICO
            fcor     = "1E4620"
            situacao = "✅ Todas as competências cobertas"

        linha = [
            empresa,
            d["cnpj"],
            colaborador,
            d["pis"],
            d["data_admissao"],
            d["data_demissao"] or "Ativo",
            len(d["esperadas"]),
            len(d["depositadas"]),
            len(d["faltantes"]),
            ", ".join(d["faltantes"]) if d["faltantes"] else "—",
            situacao,
        ]

        for c, val in enumerate(linha, 1):
            cell = ws_falt.cell(row=idx, column=c, value=val)
            align = "right" if c in (7, 8, 9) else ("center" if c in (5, 6) else "left")
            bold_flag = c in (9, 11)
            estilo_celula(cell, cor_row, bold=bold_flag, align=align,
                          font_color=fcor if tem_faltante else "000000")

    auto_width(ws_falt)
    # Coluna "Competências Faltantes" pode ser longa — limita e ativa wrap
    ws_falt.column_dimensions[get_column_letter(10)].width = 60
    for row in ws_falt.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Altura mínima para linhas com muitas faltantes
    for row_idx in range(2, len(chaves_ordenadas) + 2):
        ws_falt.row_dimensions[row_idx].height = 30

    # ── ABA SINTÉTICA ─────────────────────────────────────────────────────────
    ws_sin = wb.create_sheet("Sintético")
    ws_sin.freeze_panes = "A2"

    resumo = defaultdict(lambda: {
        "total": 0.0, "qtd": 0,
        "total_atraso": 0.0, "qtd_atraso": 0,
        "total_regular": 0.0, "qtd_regular": 0,
        "competencias": set(),
        "data_admissao": "", "data_demissao": "", "cnpj": "", "pis": "",
    })

    for r in registros:
        chave = (r["nome_empresa"], r["nome_trabalhador"])
        d = resumo[chave]
        d["total"] += r["valor"]; d["qtd"] += 1
        d["cnpj"] = r["cnpj"]; d["pis"] = r["pis"]
        d["data_admissao"] = r["data_admissao"]
        d["data_demissao"] = r["data_demissao"]
        d["competencias"].add(r["competencia"])
        if r["em_atraso"]:
            d["total_atraso"] += r["valor"]; d["qtd_atraso"] += 1
        else:
            d["total_regular"] += r["valor"]; d["qtd_regular"] += 1

    sin_headers = [
        "Empresa", "CNPJ", "Colaborador", "PIS/PASEP",
        "Data Admissão", "Data Demissão",
        "Qtd Depósitos", "Total Depositado (R$)",
        "Qtd Regulares", "Total Regular (R$)",
        "Qtd Em Atraso", "Total Em Atraso (R$)",
        "Competências Esperadas", "Competências Depositadas", "Competências Faltantes",
        "Alerta",
    ]
    for c, h in enumerate(sin_headers, 1):
        estilo_header(ws_sin.cell(row=1, column=c, value=h), COR_SIN_HEADER)
    ws_sin.row_dimensions[1].height = 28

    for idx, ((empresa, colaborador), d) in enumerate(sorted(resumo.items()), 2):
        chave = (empresa, colaborador)
        fa = analise_faltantes.get(chave, {})
        qtd_esperadas   = len(fa.get("esperadas", []))
        qtd_depositadas = len(fa.get("depositadas", set()))
        qtd_faltantes   = len(fa.get("faltantes", []))

        tem_atraso    = d["qtd_atraso"] > 0
        tem_faltante  = qtd_faltantes > 0

        # Prioridade de cor: faltante > atraso > regular
        if tem_faltante:
            cor_row = COR_FALTANTE
        elif tem_atraso:
            cor_row = COR_ATRASO
        else:
            cor_row = COR_SINTETICO

        alertas = []
        if tem_faltante:
            alertas.append(f"❌ {qtd_faltantes} comp. faltante(s)")
        if tem_atraso:
            alertas.append(f"⚠️ {d['qtd_atraso']} depósito(s) em atraso")
        if not alertas:
            alertas.append("✅ Regular")

        linha = [
            empresa, d["cnpj"], colaborador, d["pis"],
            d["data_admissao"], d["data_demissao"] or "Ativo",
            d["qtd"], d["total"],
            d["qtd_regular"], d["total_regular"],
            d["qtd_atraso"], d["total_atraso"],
            qtd_esperadas, qtd_depositadas, qtd_faltantes,
            " | ".join(alertas),
        ]

        for c, val in enumerate(linha, 1):
            cell = ws_sin.cell(row=idx, column=c, value=val)
            align = "right" if c in (7, 8, 9, 10, 11, 12, 13, 14, 15) else (
                "center" if c in (5, 6) else "left"
            )
            bold_flag = c in (15, 16) and (tem_faltante or tem_atraso)
            font_color = COR_FALTANTE_FNT if tem_faltante else "000000"
            estilo_celula(cell, cor_row, bold=bold_flag, align=align, font_color=font_color)
            if c in (8, 10, 12):
                cell.number_format = '#,##0.00'

    auto_width(ws_sin)

    # ── ABA RESUMO GERAL ──────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resumo Geral")

    total_geral   = sum(r["valor"] for r in registros)
    total_atraso  = sum(r["valor"] for r in registros if r["em_atraso"])
    total_regular = total_geral - total_atraso
    qtd_total     = len(registros)
    qtd_atraso    = sum(1 for r in registros if r["em_atraso"])
    colaboradores = len(set(r["nome_trabalhador"] for r in registros))
    empresas      = len(set(r["nome_empresa"] for r in registros))

    colab_com_faltante = sum(1 for d in analise_faltantes.values() if d["faltantes"])
    total_faltantes    = sum(len(d["faltantes"]) for d in analise_faltantes.values())
    total_esperadas    = sum(len(d["esperadas"]) for d in analise_faltantes.values())
    total_depositadas  = sum(len(d["depositadas"]) for d in analise_faltantes.values())

    linhas_res = [
        ("📊 RESUMO GERAL DA AUDITORIA FGTS", ""),
        ("Data de processamento", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Referência de competência (ativos até)", DATA_EXTRATO.strftime("%m/%Y")),
        ("", ""),
        ("EMPRESAS", ""),
        ("Qtd de Empresas", empresas),
        ("", ""),
        ("COLABORADORES", ""),
        ("Qtd de Colaboradores", colaboradores),
        ("  → Com depósitos em atraso", sum(1 for (_, colab) in
            set((r["nome_empresa"], r["nome_trabalhador"]) for r in registros if r["em_atraso"]))),
        ("  → Com competências faltantes", colab_com_faltante),
        ("", ""),
        ("DEPÓSITOS", ""),
        ("Total de Depósitos",      qtd_total),
        ("  → Regulares",           qtd_total - qtd_atraso),
        ("  → Em Atraso",           qtd_atraso),
        ("", ""),
        ("VALORES (R$)", ""),
        ("Total Geral Depositado",  total_geral),
        ("  → Total Regular",       total_regular),
        ("  → Total Em Atraso",     total_atraso),
        ("", ""),
        ("COMPETÊNCIAS", ""),
        ("Total Esperadas",         total_esperadas),
        ("  → Depositadas",         total_depositadas),
        ("  → FALTANTES ❌",        total_faltantes),
        ("", ""),
        ("⚠️  ATENÇÃO – DEPÓSITOS EM ATRASO",
         "O valor pode divergir dos 8% devidos sobre o salário bruto da competência original."),
        ("❌  ATENÇÃO – COMPETÊNCIAS FALTANTES",
         "Meses em que nenhum depósito (regular ou retroativo) foi identificado no extrato. "
         "Consultar aba 'Competências Faltantes' para detalhe por colaborador."),
    ]

    ws_res.column_dimensions["A"].width = 42
    ws_res.column_dimensions["B"].width = 80

    for row_i, (label, valor) in enumerate(linhas_res, 1):
        c_label = ws_res.cell(row=row_i, column=1, value=label)
        c_valor = ws_res.cell(row=row_i, column=2, value=valor)

        if label.startswith("📊") or (label.isupper() and label.strip()):
            c_label.font = Font(bold=True, size=11, color="1F3864")
            for c in (c_label, c_valor):
                c.fill = PatternFill("solid", start_color="D9E1F2")
        elif label.startswith("  →"):
            for c in (c_label, c_valor):
                c.font = Font(italic=True, size=9, color="404040")
        elif "ATENÇÃO" in label and "FALTANTES" in label:
            for c in (c_label, c_valor):
                c.font = Font(bold=True if c == c_label else False, color=COR_FALTANTE_FNT)
                c.fill = PatternFill("solid", start_color=COR_FALTANTE)
        elif "ATENÇÃO" in label:
            for c in (c_label, c_valor):
                c.font = Font(bold=True if c == c_label else False, color="7F4D00")
                c.fill = PatternFill("solid", start_color="FFF2CC")

        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            c_valor.alignment = Alignment(horizontal="right")
            if isinstance(valor, float):
                c_valor.number_format = 'R$ #,##0.00'

    wb.save(caminho_saida)
    print(f"\n✅ Relatório gerado com sucesso: {caminho_saida}")

    caminho_html = os.path.splitext(caminho_saida)[0] + ".html"
    gerar_dashboard_html(registros, analise_faltantes, caminho_html)

    return {
        "empresas": empresas,
        "colaboradores": colaboradores,
        "total_depositos": qtd_total,
        "qtd_regular": qtd_total - qtd_atraso,
        "qtd_atraso": qtd_atraso,
        "total_geral": total_geral,
        "total_regular": total_regular,
        "total_atraso": total_atraso,
        "total_esperadas": total_esperadas,
        "total_depositadas": total_depositadas,
        "total_faltantes": total_faltantes,
        "colab_com_faltante": colab_com_faltante,
    }


# ──────────────────────────────────────────────────────────────────────────────
# DASHBOARD HTML NARRATIVO
# ──────────────────────────────────────────────────────────────────────────────

def calcular_series_anuais(registros: list[dict], analise_faltantes: dict) -> dict:
    """
    Monta todas as séries anuais necessárias para o dashboard:
    - reg[ano], atr[ano], falt[ano], colab_ativos[ano], pct_atraso[ano]
    - admissoes[ano], demissoes[ano]
    - valor_reg[ano], valor_atr[ano]
    """
    from dateutil.relativedelta import relativedelta

    dep_por_colab: dict[str, set] = defaultdict(set)
    for r in registros:
        dep_por_colab[r["nome_trabalhador"]].add(
            (int(r["competencia"].split("/")[1]), int(r["competencia"].split("/")[0]))
        )

    stats_dep: dict[int, dict] = defaultdict(lambda: {"reg": 0, "atr": 0, "v_reg": 0.0, "v_atr": 0.0})
    for r in registros:
        try:
            ano = int(r["competencia"].split("/")[1])
        except Exception:
            continue
        if r["em_atraso"]:
            stats_dep[ano]["atr"] += 1
            stats_dep[ano]["v_atr"] += r["valor"]
        else:
            stats_dep[ano]["reg"] += 1
            stats_dep[ano]["v_reg"] += r["valor"]

    # Faltantes por ano de competência
    faltantes_ano: dict[int, int] = defaultdict(int)
    colab_ativos_ano: dict[int, set] = defaultdict(set)
    admissoes_ano: dict[int, int] = defaultdict(int)
    demissoes_ano: dict[int, int] = defaultdict(int)

    for (_, colab), d in analise_faltantes.items():
        adm = d["data_admissao"]
        dem = d["data_demissao"] or None
        if not adm:
            continue
        try:
            da, ma, aa = adm.split("/")
            inicio = date(int(aa), int(ma), int(da))
            admissoes_ano[int(aa)] += 1
        except Exception:
            continue
        if dem:
            try:
                dd2, md2, ad2 = dem.split("/")
                fim = date(int(ad2), int(md2), int(dd2))
                demissoes_ano[int(ad2)] += 1
            except Exception:
                fim = DATA_EXTRATO.replace(day=1) - relativedelta(months=1)
        else:
            fim = DATA_EXTRATO.replace(day=1) - relativedelta(months=1)

        cursor = date(inicio.year, inicio.month, 1)
        while cursor <= fim:
            colab_ativos_ano[cursor.year].add(colab)
            if (cursor.year, cursor.month) not in dep_por_colab[colab]:
                faltantes_ano[cursor.year] += 1
            cursor += relativedelta(months=1)

    todos_anos = sorted(
        set(list(stats_dep.keys()) + list(faltantes_ano.keys()) + list(colab_ativos_ano.keys()))
    )

    result = {}
    for a in todos_anos:
        s = stats_dep[a]
        tot = s["reg"] + s["atr"]
        result[a] = {
            "reg":    s["reg"],
            "atr":    s["atr"],
            "v_reg":  round(s["v_reg"], 2),
            "v_atr":  round(s["v_atr"], 2),
            "falt":   faltantes_ano[a],
            "colab":  len(colab_ativos_ano[a]),
            "pct":    round(100 * s["atr"] / tot, 1) if tot else 0.0,
            "adm":    admissoes_ano[a],
            "dem":    demissoes_ano[a],
        }
    return result


def _narrativa(series: dict) -> list[tuple[str, str, str]]:
    """
    Gera automaticamente os marcos narrativos baseados nos dados reais.
    Retorna lista de (ano_label, tipo [red/amber/green/gray], texto).
    """
    anos = sorted(series.keys())
    if not anos:
        return []

    eventos = []

    # Agrupa anos consecutivos com comportamento similar
    def grupo(anos_lista, fn_tipo):
        grupos = []
        atual = [anos_lista[0]]
        for a in anos_lista[1:]:
            if fn_tipo(a) == fn_tipo(atual[-1]):
                atual.append(a)
            else:
                grupos.append(atual)
                atual = [a]
        grupos.append(atual)
        return grupos

    # Anos sem nenhum depósito mas com colaboradores ativos
    sem_dep = [a for a in anos if series[a]["reg"] + series[a]["atr"] == 0 and series[a]["colab"] > 0]
    if sem_dep:
        inicio = sem_dep[0]; fim = sem_dep[-1]
        label = str(inicio) if inicio == fim else f"{inicio} – {fim}"
        total_falt = sum(series[a]["falt"] for a in sem_dep)
        max_colab = max(series[a]["colab"] for a in sem_dep)
        eventos.append((inicio, label, "gray",
            f"Nenhum depósito de FGTS realizado, com até {max_colab} colaborador(es) ativo(s) no período. "
            f"{total_falt} competências faltantes acumuladas silenciosamente."))

    # Primeiro depósito
    primeiro_dep = next((a for a in anos if series[a]["reg"] + series[a]["atr"] > 0), None)
    if primeiro_dep:
        s = series[primeiro_dep]
        pct = s["pct"]
        tipo = "red" if pct >= 80 else "amber" if pct >= 30 else "green"
        eventos.append((primeiro_dep, str(primeiro_dep), tipo,
            f"Primeiros depósitos registrados — mas {pct}% já em atraso "
            f"({s['atr']} de {s['reg']+s['atr']}). O passivo retroativo começa a aparecer."))

    # Anos de 100% inadimplência (exceto o primeiro)
    cem_pct = [a for a in anos if series[a]["pct"] == 100.0 and a != primeiro_dep]
    if cem_pct:
        grupos_100 = []
        atual = [cem_pct[0]]
        for a in cem_pct[1:]:
            if a == atual[-1] + 1:
                atual.append(a)
            else:
                grupos_100.append(atual)
                atual = [a]
        grupos_100.append(atual)
        for g in grupos_100:
            label = str(g[0]) if len(g) == 1 else f"{g[0]} – {g[-1]}"
            total_atr = sum(series[a]["atr"] for a in g)
            max_colab = max(series[a]["colab"] for a in g)
            total_falt = sum(series[a]["falt"] for a in g)
            eventos.append((g[0], label, "red",
                f"{len(g)} ano(s) consecutivo(s) com 100% de inadimplência — {total_atr} depósitos, "
                f"todos em atraso. Quadro de até {max_colab} colaborador(es) ativo(s). "
                f"{total_falt} novas competências faltantes acumuladas."))

    # Anos de melhora parcial (pct entre 10% e 79%)
    melhora = [a for a in anos if 0 < series[a]["pct"] < 80 and a != primeiro_dep and series[a]["reg"] > 0]
    for a in melhora:
        s = series[a]
        eventos.append((a, str(a), "amber",
            f"Sinal de regularização parcial: {s['reg']} depósito(s) regular(es) e {s['atr']} em atraso "
            f"({s['pct']}% de atraso). O esforço não foi suficiente para cobrir todas as competências."))

    # Anos regulares (pct = 0 e há depósitos)
    regulares = [a for a in anos if series[a]["pct"] == 0 and series[a]["reg"] > 0]
    if regulares:
        label = str(regulares[0]) if len(regulares) == 1 else f"{regulares[0]}–{regulares[-1]}"
        eventos.append((regulares[0], label, "green",
            f"Depósitos realizados em dia ({sum(series[a]['reg'] for a in regulares)} no total). "
            "Período de conformidade registrado."))

    # Anos recentes sem depósito com quadro ativo (possível paralisação)
    ano_corte = max(anos) - 2
    recentes_sem = [a for a in anos if a >= ano_corte and series[a]["reg"] + series[a]["atr"] == 0 and series[a]["colab"] > 0]
    if recentes_sem:
        total_falt = sum(series[a]["falt"] for a in recentes_sem)
        max_colab = max(series[a]["colab"] for a in recentes_sem)
        label = str(recentes_sem[0]) if len(recentes_sem) == 1 else f"{recentes_sem[0]} – {recentes_sem[-1]}"
        eventos.append((recentes_sem[0] + 0.5, label, "gray",
            f"Nenhum depósito identificado nos anos mais recentes, com {max_colab} colaborador(es) ainda ativo(s). "
            f"{total_falt} novas competências faltantes. Possível paralisação total dos depósitos."))

    eventos.sort(key=lambda x: x[0])
    return [(label, tipo, texto) for (_, label, tipo, texto) in eventos]


def gerar_dashboard_html(registros: list[dict], analise_faltantes: dict, caminho_html: str):
    """Gera dashboard HTML narrativo standalone."""

    series = calcular_series_anuais(registros, analise_faltantes)
    narrativa = _narrativa(series)

    anos = sorted(series.keys())
    js_anos  = str(anos)
    js_reg   = str([series[a]["reg"]   for a in anos])
    js_atr   = str([series[a]["atr"]   for a in anos])
    js_falt  = str([series[a]["falt"]  for a in anos])
    js_pct   = str([series[a]["pct"]   for a in anos])
    js_colab = str([series[a]["colab"] for a in anos])

    # KPIs gerais
    total_dep   = len(registros)
    qtd_atr     = sum(1 for r in registros if r["em_atraso"])
    pct_atr_geral = round(100 * qtd_atr / total_dep, 1) if total_dep else 0
    total_val   = sum(r["valor"] for r in registros)
    total_falt  = sum(len(d["faltantes"]) for d in analise_faltantes.values())
    colab_falt  = sum(1 for d in analise_faltantes.values() if d["faltantes"])
    total_colab = len(analise_faltantes)
    empresa_nome = registros[0]["nome_empresa"] if registros else "Empresa"
    cnpj         = registros[0]["cnpj"] if registros else ""
    data_proc    = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Gera HTML dos marcos narrativos
    tl_html = ""
    for label, tipo, texto in narrativa:
        cor_dot = {"red": "#E24B4A", "amber": "#EF9F27", "green": "#639922", "gray": "#888780"}[tipo]
        cor_bdr = {"red": "#A32D2D", "amber": "#854F0B", "green": "#3B6D11", "gray": "#5F5E5A"}[tipo]
        cor_bg  = {"red": "#FCEBEB", "amber": "#FAEEDA", "green": "#EAF3DE", "gray": "#F1EFE8"}[tipo]
        tl_html += f"""
        <div class="tl-item">
          <span class="tl-dot" style="background:{cor_dot};border-color:{cor_bdr}"></span>
          <span class="tl-year">{label}</span>
          <span class="tl-text">{texto}</span>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Auditoria FGTS — {empresa_nome}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f5f4f0;color:#1a1a18;font-size:14px;line-height:1.6}}
  .page{{max-width:960px;margin:0 auto;padding:32px 24px 64px}}
  .header{{margin-bottom:28px;padding-bottom:20px;border-bottom:1.5px solid #d3d1c7}}
  .header-top{{display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px}}
  .company-name{{font-size:22px;font-weight:600;color:#1a1a18;margin-bottom:3px}}
  .company-meta{{font-size:12px;color:#5f5e5a}}
  .badge-proc{{font-size:11px;color:#5f5e5a;background:#e8e7e1;padding:4px 10px;border-radius:4px}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:28px}}
  .kpi{{background:#fff;border-radius:8px;border:0.5px solid #d3d1c7;padding:16px}}
  .kpi-label{{font-size:11px;color:#888780;margin-bottom:4px;text-transform:uppercase;letter-spacing:.04em}}
  .kpi-value{{font-size:24px;font-weight:600}}
  .kpi-value.danger{{color:#A32D2D}}
  .kpi-value.warn{{color:#854F0B}}
  .kpi-value.ok{{color:#3B6D11}}
  .section-label{{font-size:11px;font-weight:600;color:#888780;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px}}
  .chart-block{{background:#fff;border-radius:8px;border:0.5px solid #d3d1c7;padding:20px;margin-bottom:16px}}
  .chart-wrap{{position:relative;width:100%}}
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
  .legend-row{{display:flex;flex-wrap:wrap;gap:14px;margin-bottom:12px;font-size:12px;color:#5f5e5a}}
  .leg{{display:flex;align-items:center;gap:5px}}
  .leg-sq{{width:10px;height:10px;border-radius:2px;flex-shrink:0}}
  .narrative-block{{background:#fff;border-radius:8px;border:0.5px solid #d3d1c7;padding:20px;margin-bottom:16px}}
  .timeline{{padding-left:22px;position:relative}}
  .timeline::before{{content:'';position:absolute;left:7px;top:0;bottom:0;width:1.5px;background:#d3d1c7}}
  .tl-item{{position:relative;margin-bottom:18px}}
  .tl-dot{{position:absolute;left:-18px;top:4px;width:11px;height:11px;border-radius:50%;border:2px solid}}
  .tl-year{{display:block;font-size:11px;font-weight:600;color:#5f5e5a;margin-bottom:3px;letter-spacing:.03em}}
  .tl-text{{font-size:13px;color:#2c2c2a;line-height:1.55}}
  .footer{{font-size:11px;color:#888780;margin-top:32px;text-align:center;border-top:0.5px solid #d3d1c7;padding-top:16px}}
  @media(max-width:600px){{.kpi-grid{{grid-template-columns:repeat(2,1fr)}}.two-col{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<div class="page">

  <div class="header">
    <div class="header-top">
      <div>
        <p class="company-name">{empresa_nome}</p>
        <p class="company-meta">CNPJ {cnpj} &nbsp;·&nbsp; Extrato analisado: {anos[0] if anos else '—'} – {anos[-1] if anos else '—'}</p>
      </div>
      <span class="badge-proc">Processado em {data_proc}</span>
    </div>
  </div>

  <div class="kpi-grid">
    <div class="kpi">
      <p class="kpi-label">Total depositado</p>
      <p class="kpi-value">R$ {total_val:,.0f}</p>
    </div>
    <div class="kpi">
      <p class="kpi-label">Depósitos em atraso</p>
      <p class="kpi-value danger">{pct_atr_geral}%</p>
    </div>
    <div class="kpi">
      <p class="kpi-label">Competências faltantes</p>
      <p class="kpi-value danger">{total_falt:,}</p>
    </div>
    <div class="kpi">
      <p class="kpi-label">Colaboradores afetados</p>
      <p class="kpi-value warn">{colab_falt} / {total_colab}</p>
    </div>
  </div>

  <div class="chart-block">
    <p class="section-label">Conformidade por ano de competência</p>
    <div class="legend-row">
      <span class="leg"><span class="leg-sq" style="background:#639922"></span>Regulares</span>
      <span class="leg"><span class="leg-sq" style="background:#EF9F27"></span>Em atraso</span>
      <span class="leg"><span class="leg-sq" style="background:#E24B4A"></span>Faltantes</span>
    </div>
    <div class="chart-wrap" style="height:220px">
      <canvas id="chartStack" role="img" aria-label="Gráfico empilhado de depósitos regulares, em atraso e faltantes por ano">Dados de conformidade FGTS por ano.</canvas>
    </div>
  </div>

  <div class="two-col">
    <div class="chart-block">
      <p class="section-label">% de atraso por ano</p>
      <div class="chart-wrap" style="height:180px">
        <canvas id="chartPct" role="img" aria-label="Percentual de depósitos em atraso por ano">Percentual de atraso por ano.</canvas>
      </div>
    </div>
    <div class="chart-block">
      <p class="section-label">Colaboradores ativos vs faltantes</p>
      <div class="legend-row">
        <span class="leg"><span class="leg-sq" style="background:#378ADD"></span>Ativos</span>
        <span class="leg"><span class="leg-sq" style="background:#E24B4A;opacity:.6"></span>Faltantes (eixo dir.)</span>
      </div>
      <div class="chart-wrap" style="height:160px">
        <canvas id="chartColab" role="img" aria-label="Colaboradores ativos versus competências faltantes por ano">Colaboradores ativos e faltantes por ano.</canvas>
      </div>
    </div>
  </div>

  <div class="narrative-block">
    <p class="section-label" style="margin-bottom:16px">A história desta empresa</p>
    <div class="timeline">
      {tl_html}
    </div>
  </div>

  <div class="footer">
    Auditoria FGTS — gerado por auditor_fgts.py &nbsp;·&nbsp; {data_proc}
  </div>

</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const anos  = {js_anos};
const reg   = {js_reg};
const atr   = {js_atr};
const falt  = {js_falt};
const pct   = {js_pct};
const colab = {js_colab};

const C_GREEN='#639922',C_AMBER='#EF9F27',C_RED='#E24B4A',C_BLUE='#378ADD';
const gridCol='rgba(0,0,0,0.07)', tickCol='#888780';
Chart.defaults.font.family='-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif';
Chart.defaults.font.size=11;

new Chart(document.getElementById('chartStack'),{{
  type:'bar',
  data:{{labels:anos,datasets:[
    {{label:'Regulares',data:reg,backgroundColor:C_GREEN,stack:'s'}},
    {{label:'Em atraso',data:atr,backgroundColor:C_AMBER,stack:'s'}},
    {{label:'Faltantes',data:falt,backgroundColor:C_RED,stack:'s'}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{
      x:{{stacked:true,ticks:{{color:tickCol,autoSkip:false,maxRotation:45}},grid:{{color:gridCol}}}},
      y:{{stacked:true,ticks:{{color:tickCol}},grid:{{color:gridCol}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartPct'),{{
  type:'line',
  data:{{labels:anos,datasets:[{{
    label:'% em atraso',data:pct,
    borderColor:C_RED,backgroundColor:'rgba(226,75,74,0.1)',
    fill:true,tension:0.3,pointRadius:4,pointBackgroundColor:C_RED
  }}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{
      x:{{ticks:{{color:tickCol,autoSkip:true,maxRotation:45}},grid:{{color:gridCol}}}},
      y:{{min:0,max:110,ticks:{{color:tickCol,callback:v=>v+'%'}},grid:{{color:gridCol}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartColab'),{{
  type:'bar',
  data:{{labels:anos,datasets:[
    {{label:'Colaboradores ativos',data:colab,backgroundColor:'rgba(55,138,221,0.4)',borderColor:C_BLUE,borderWidth:1,yAxisID:'y'}},
    {{label:'Comp. faltantes',data:falt,type:'line',borderColor:C_RED,backgroundColor:'transparent',
      tension:0.3,pointRadius:3,pointBackgroundColor:C_RED,borderDash:[4,3],yAxisID:'y2'}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{
      x:{{ticks:{{color:tickCol,autoSkip:true,maxRotation:45}},grid:{{color:gridCol}}}},
      y:{{ticks:{{color:C_BLUE}},grid:{{color:gridCol}}}},
      y2:{{position:'right',ticks:{{color:C_RED}},grid:{{drawOnChartArea:false}}}}
    }}
  }}
}});
</script>
</body>
</html>
"""
    with open(caminho_html, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"   Dashboard HTML  : {caminho_html}")


# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python auditor_fgts.py <entrada.xlsx> [saida.xlsx]")
        sys.exit(1)

    entrada = sys.argv[1]
    if not os.path.exists(entrada):
        print(f"❌ Arquivo não encontrado: {entrada}")
        sys.exit(1)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    saida = sys.argv[2] if len(sys.argv) >= 3 else f"auditoria_fgts_{ts}.xlsx"

    print(f"📂 Lendo: {entrada}")
    registros = parse_extrato(entrada)
    print(f"   {len(registros)} depósitos FGTS encontrados")

    stats = gerar_xlsx(registros, saida)

    print("\n── Depósitos ───────────────────────────────────────")
    print(f"   Empresas          : {stats['empresas']}")
    print(f"   Colaboradores     : {stats['colaboradores']}")
    print(f"   Total depósitos   : {stats['total_depositos']}")
    print(f"   Regulares         : {stats['qtd_regular']}")
    print(f"   Em atraso ⚠️       : {stats['qtd_atraso']}")
    print(f"   Valor total       : R$ {stats['total_geral']:,.2f}")
    print(f"   Valor regular     : R$ {stats['total_regular']:,.2f}")
    print(f"   Valor em atraso   : R$ {stats['total_atraso']:,.2f}")
    print("\n── Competências ────────────────────────────────────")
    print(f"   Esperadas         : {stats['total_esperadas']}")
    print(f"   Depositadas       : {stats['total_depositadas']}")
    print(f"   Faltantes ❌       : {stats['total_faltantes']}")
    print(f"   Colab. c/ faltante: {stats['colab_com_faltante']}")
    print("────────────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()
